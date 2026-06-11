"""Excel (.xlsx) parser with section_ref support."""

from pathlib import Path

from .base import BaseParser, ParseResult, Section


class XLSXParser(BaseParser):
    """Parser for .xlsx files using openpyxl."""

    def parse(self, file_path: Path) -> ParseResult:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sections: list[Section] = []
        total_rows = 0
        all_text_parts: list[str] = []

        for idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                sections.append(Section(
                    title=sheet_name,
                    content="(empty sheet)",
                    section_ref=f"sheet-{idx}",
                ))
                continue

            total_rows += len(rows)

            # First row as header
            headers = [str(c) if c is not None else "" for c in rows[0]]
            data_rows = rows[1:]

            lines: list[str] = []
            lines.append(" | ".join(headers))
            lines.append(" | ".join("---" for _ in headers))
            for row in data_rows:
                cells = [str(c) if c is not None else "" for c in row]
                lines.append(" | ".join(cells))

            content = "\n".join(lines)
            sections.append(Section(
                title=sheet_name,
                content=content,
                section_ref=f"sheet-{idx}",
            ))
            all_text_parts.append(f"=== {sheet_name} ===\n{content}")

        wb.close()

        return ParseResult(
            filename=file_path.name,
            sections=sections,
            raw_text="\n\n".join(all_text_parts),
            page_count=len(wb.sheetnames),
            metadata={
                "parser": "openpyxl",
                "sheet_count": len(wb.sheetnames),
                "total_rows": total_rows,
                "sheet_names": list(wb.sheetnames),
            },
        )
